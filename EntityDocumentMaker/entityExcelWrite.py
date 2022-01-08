# -*- coding: utf-8 -*-
import sys, codecs
import os
import db

import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import PatternFill
from sqlalchemy import and_, or_, not_

from models import EntityMaster, ColumnMaster
from SettingEntityAccess import SettingEntityAccess

from ConfigManager import ConfigManager

from models import *

# エンティティ定義書書き込みクラス
class EntityExcelWrite(): #クラス名

    # インスタンス化メソッド
    def __init__(self, del_flg_skp):
        configManager = ConfigManager()
        # Entity Document Template Path
        self.template_file_path = r'{}'.format(configManager.getEntityDocTemplatePath())
        self.tamplate_sheet_name = configManager.getEntitySheetName()
        self.revision_history_sheet_name = configManager.getRevisionHistorySheetName()

        self.del_flg_skp = del_flg_skp

    # Entity全選択メソッド
    def entityAllWrite(self, full_path):
        # 設定マスタを全取得
        entityMasters = db.session.query(EntityMaster).\
                            filter(and_(EntityMaster.entity_no == entity_no, EntityMaster.entity_cd == entity_cd)).\
                            order_by(EntityMaster.id).all()

        #　データベース.セッションをクローズ
        db.session.close()

        for entityMaster in entityMasters:
            self.entityWrite(full_path=full_path,
                            entity_no=entityMaster.entity_no,
                            entity_cd=entityMaster.entity_cd,
                            entity_name=entityMaster.entity_name,
                            type_section_name=entityMaster.type_section_name)

    # Entity選択メソッド
    def entitySelecttWrite(self, full_path, entity_selects):

        # 辞書ループ
        for key,value in entity_selects.items():

            entity_no = key
            entity_cd = value

            # 設定マスタを全取得
            entityMasters = db.session.query(EntityMaster).\
                                filter(and_(EntityMaster.entity_no == entity_no, EntityMaster.entity_cd == entity_cd)).\
                                order_by(EntityMaster.id).all()

            #　データベース.セッションをクローズ
            # db.session.close()

            for entityMaster in entityMasters:
                self.entityWrite(full_path=full_path,
                                entity_no=entityMaster.entity_no,
                                entity_cd=entityMaster.entity_cd,
                                entity_name=entityMaster.entity_name,
                                type_section_name=entityMaster.type_section_name)

    # メイン処理
    def entityWrite(self, full_path, entity_no, entity_cd, entity_name, type_section_name):
        path = SQLITE3_NAME
        if not os.path.isfile(path):

            # テーブルを作成する
            Base.metadata.create_all(db.engine)

        # エンティティ定義書読み込みファイル設定
        wb = openpyxl.load_workbook(self.template_file_path)
        ws = wb[self.tamplate_sheet_name]

        ws['A2'] = entity_name
        ws['A4'] = entity_cd

        # 設定マスタを全取得
        columnMasters = db.session.query(ColumnMaster).\
                            filter(and_(ColumnMaster.entity_no == entity_no, ColumnMaster.entity_cd == entity_cd)).\
                            order_by(ColumnMaster.id).all()

        g0625_fill = PatternFill(
                    fgColor='C0C0C0', bgColor='C0C0C0', fill_type='gray0625')

        start_row = 7
        for index,columnMaster in enumerate(columnMasters):
            row_count = start_row + index

            not_null=''
            if columnMaster.not_null_flg > 0:
                not_null='〇'

            primary_key=''
            if columnMaster.primary_key_flg > 0:
                primary_key='〇'

            # Column No
            ws['A{}'.format(row_count)] = columnMaster.column_no
            # Column Cd
            ws['B{}'.format(row_count)] = columnMaster.column_cd
            # Column名
            ws['C{}'.format(row_count)] = columnMaster.column_name
            # 型
            ws['D{}'.format(row_count)] = columnMaster.modl
            # 桁
            ws['E{}'.format(row_count)] = columnMaster.digit
            # 制度
            ws['F{}'.format(row_count)] = columnMaster.accuracy
            # Not Null 制約
            ws['G{}'.format(row_count)] = not_null
            # primary
            ws['H{}'.format(row_count)] = primary_key
            # 説明 H,I,J
            ws['I{}'.format(row_count)] = columnMaster.explanation
            # DB制約
            ws['L{}'.format(row_count)] = columnMaster.db_constraint
            # 備考
            ws['M{}'.format(row_count)] = columnMaster.remarks
            # ws['N{}'.format(row_count)] = columnMaster.entity_cd

            # セル背景色削除
            if columnMaster.del_flg > 0:
                for row in ws['A{}:M{}'.format(row_count, row_count)]:
                    for cell in row:
                        cell.fill = g0625_fill

                ws['N{}'.format(row_count)] = '削除'


        # 保存する
        wb.save(r'{}\No{}_{}{}'.format(full_path, f'{entity_no:03}', entity_name, '.xlsx'))

# クラス開始処理
#　Pythonファイル直接起動時に呼び出される
if __name__ == '__main__':

    full_path = r'C:\Entity_Document_Search\output'

    entity_no = 7
    entity_cd = 'SUPPLIERS_MASTER'

    entity_selects = { entity_no : entity_cd }

    entity_no = 6
    entity_cd = 'USER_MASTER'

    entity_selects.update({entity_no:entity_cd})

    # エンティティ定義書書き込み処理
    entityWrite = EntityExcelWrite(0)
    entityWrite.entitySelecttWrite(full_path=full_path, entity_selects=entity_selects)
