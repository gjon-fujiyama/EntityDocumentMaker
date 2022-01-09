#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys, codecs
import os
import db

import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import PatternFill

from models import EntityMaster, ColumnMaster
from SettingEntityAccess import SettingEntityAccess
from ConfigManager import ConfigManager

# エンティティ定義書読み取りクラス
class EntityRead(): #クラス名

    # インスタンス化メソッド
    def __init__(self):
        self.configManager = ConfigManager()
        self.target_extension = self.configManager.getTargetExtension()
        self.target_filename_prefix = self.configManager.getTargetFilenamePrefix()

    # メイン処理
    def startRead(self):

        sea = SettingEntityAccess()
        settingEntitys = sea.getAllEntityPath()
        print(list(settingEntitys))

        db.session.query(EntityMaster).delete()  # 削除
        db.session.query(ColumnMaster).delete()  # 削除

        # 設定リストループ
        entityNo = 1
        for settingEntity in settingEntitys:

            path = settingEntity.set_import_path

            # 対象になるエンティティ定義書フォルダ配下の定義書ファイル名を全取得
            files = os.listdir(path)
            files_file = [f for f in files if os.path.isfile(os.path.join(path, f))]
            print(files_file)

            # 取得ファイル名分反復処理
            for filename in files_file:

                # 対象拡張子のみ取込み
                if self.target_extension in filename:
                    # 対象ファイル接頭辞のもののみ取込み
                    if self.target_filename_prefix in filename:

                        # フルパス構築
                        full_path = '{}\{}'.format(path, filename)

                        # print('full_path:{}'.format(full_path))

                        # エンティティ定義書読み取り処理
                        self.entityRead(full_path=full_path,
                                        entityNo=entityNo,
                                        type_section=settingEntity.type_section,
                                        type_section_name=settingEntity.type_section_name)
                        entityNo += 1

        return entityNo;

    # メイン処理
    def entityRead(self, full_path, entityNo, type_section, type_section_name):

        # エンティティ定義書読み込みファイル設定
        wb = openpyxl.load_workbook(full_path)
        ws = wb['エンティティ定義書']

        # エンティティ定義項目最大値行
        maxCount = self.maxRowCount(ws) + 1
        # 現在日時取得
        now = datetime.datetime.now()

        entity_cd = ""
        if ws['A4'].value != None:
            entity_cd=ws['A4'].value

        entity_name = ""
        if ws['A2'].value != None:
            entity_name = ws['A2'].value

        # エンティティマスタ 説明　値存在判定
        explanation = ""
        if ws['A5'].value != None:
            explanation = ws['A5'].value

        # Table Section Key
        entity_section = self.configManager.getTableTypeSectionValue('Table')

        # エンティティマスタ設定
        entityMaster = EntityMaster(entity_cd=entity_cd,
                                    entity_name=entity_name,
                                    entity_no=entityNo,
                                    explanation=explanation,
                                    entity_section=entity_section,
                                    del_flg=0,
                                    entity_document_path=full_path,
                                    type_section=type_section,
                                    type_section_name=type_section_name,
                                    register_date=now,
                                    register_cd="EntityRead",
                                    update_date=now,
                                    update_cd="EntityRead")

        # エンティティマスタ　データ登録
        db.session.add(entityMaster)  # 追加
        db.session.commit()  # データベースコミット

        # 灰色系統のRGB配列（標準のカラーにある灰色１５色を想定）
        color_list = ['FFF8F8F8',
                    'FFEAEAEA',
                    'FFDDDDDD',
                    'FFC0C0C0',
                    'FFB2B2B2',
                    'FF969696',
                    'FF808080',
                    'FF808080',
                    'FF5F5F5F',
                    'FF4D4D4D',
                    'FF333333',
                    'FF292929',
                    'FF1C1C1C',
                    'FF111111',
                    'FF080808']

        # カラムマスタデータ反復処理
        for num in range(7, maxCount):
            # 項目指定のセル列
            cel_addressA = 'A{}'.format(str(num))
            cel_addressB = 'B{}'.format(str(num))
            cel_addressC = 'C{}'.format(str(num))
            cel_addressD = 'D{}'.format(str(num))
            cel_addressE = 'E{}'.format(str(num))
            cel_addressF = 'F{}'.format(str(num))
            cel_addressG = 'G{}'.format(str(num))
            cel_addressH = 'H{}'.format(str(num))
            cel_addressI = 'I{}'.format(str(num))
            cel_addressL = 'L{}'.format(str(num))
            cel_addressM = 'M{}'.format(str(num))

            # セル背景情報取得
            cel_color_type = ws[cel_addressA].fill.fgColor.type
            cel_color_theme = ws[cel_addressA].fill.fgColor.theme
            cel_color_tint = ws[cel_addressA].fill.fgColor.tint
            cel_color_rgb = ws[cel_addressA].fill.fgColor.rgb

            cel_color = ""
            # 削除フラグ判定
            del_flg = 0
            if cel_color_type=='theme':
                # 配色テーマ取得debug
                cel_color = 'theme={}'.format(cel_color_theme)
                #色合い＝０「灰色系テーマ」だった場合
                if cel_color_theme==0:
                    del_flg = 1
            else:
                # 配色RGB取得debug
                cel_color = 'rgb={}'.format(cel_color_rgb)

                #　灰色系RGBを配列判定
                if cel_color_rgb in color_list:
                    del_flg = 1

            # セル背景情報出力debug
            # print('{0} type={1}, {2}, tint={3}'.format(cel_addressA, cel_color_type, cel_color, cel_color_tint))

            column_no = num
            if ws[cel_addressA].value!=None:
                column_no = int(ws[cel_addressA].value)

            column_cd = ''
            if ws[cel_addressC].value!=None:
                column_cd = ws[cel_addressC].value

            column_name = ''
            if ws[cel_addressB].value!=None:
                column_name = ws[cel_addressB].value

            modl = ''
            if ws[cel_addressD].value!=None:
                modl = ws[cel_addressD].value

            # 桁数判定
            # digit = 0
            digit = ''
            if ws[cel_addressE].value!=None:
                #  digit = int(ws[cel_addressE].value)
                digit = ws[cel_addressE].value

            # 精度判定
            # accuracy = 0
            accuracy = ''
            if ws[cel_addressF].value!=None:
                #  accuracy = int(ws[cel_addressF].value)
                accuracy = ws[cel_addressF].value

            # NOT NULL判定
            not_null_flg = 0
            if ws[cel_addressG].value=='〇':
                not_null_flg = 1

            # プライマリキー判定
            primary_key_flg = 0
            if ws[cel_addressH].value=='〇':
                primary_key_flg = 1

            # 説明判定
            explanation = ''
            if ws[cel_addressI].value!=None:
                explanation = ws[cel_addressI].value

            # ＤＢ制約判定
            db_constraint = ''
            if ws[cel_addressL].value!=None:
                db_constraint = ws[cel_addressL].value

            # 特記事項判定
            remarks = ''
            if ws[cel_addressM].value!=None:
                remarks = ws[cel_addressM].value

            # エンティティ定義　項目情報出力
            columnMaster = ColumnMaster(entity_cd=entityMaster.entity_cd,
                                        column_cd=column_cd,
                                        column_name=column_name,
                                        ref_entity_cd="",
                                        ref_entity_name="",
                                        column_no=column_no,
                                        modl=modl,
                                        digit=digit,
                                        accuracy=accuracy,
                                        not_null_flg=not_null_flg,
                                        primary_key_flg=primary_key_flg,
                                        explanation=explanation,
                                        db_constraint=db_constraint,
                                        remarks=remarks,
                                        del_flg=del_flg,
                                        entity_no=entityNo,
                                        register_date=now,
                                        register_cd="U000001",
                                        update_date=now,
                                        update_cd="U000001");

            # カラムマスター設定出力
            # print('ColumnMaster =>>> {}'.format(columnMaster))

            db.session.add(columnMaster)
            db.session.commit()

        db.session.close()  # セッションを閉じる


    # 最大行をB列を元に再算出
    def maxRowCount(self, ws):
        #最大行
        maxRow = ws.max_row + 1

        #最大列
        maxClm = ws.max_column + 1

        # 最大行から上の行へ逆ループ
        for i in reversed(range(1, maxRow)):

            #　B列のセルがNoneではなければ
            if ws.cell(row=i, column=2).value != None:
                r_maxRow = i
                #表示
                # print('最終行：' + str(i))
                break

        # 最大行返却
        return i

# クラス開始処理
#　Pythonファイル直接起動時に呼び出される
if __name__ == '__main__':

    # エンティティ定義書読み取り処理
    entityRead = EntityRead()
    entityRead.startRead()
